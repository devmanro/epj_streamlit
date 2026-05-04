import os
import time
import logging
import traceback
import io
import zipfile
import tempfile
import requests
from pathlib import Path
from typing import List, Tuple, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border
import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib import colors as mcolors
import numpy as np
import re


# ==========================================================
# ====== TOTAL/J LANDMARK + BORDERS DETECTION ==============
# ==========================================================


def cell_has_any_border(cell) -> bool:
    try:
        b = cell.border
        sides = [b.left, b.right, b.top, b.bottom]
        return any(
            s is not None and s.border_style is not None and s.border_style != "none"
            for s in sides
        )
    except Exception:
        return False


def find_totalj_col(
    sheet, start_row: int, end_row: int, start_col: int, end_col: int
) -> Optional[int]:
    pattern = re.compile(r"total\s*/\s*j", re.IGNORECASE)
    for col in range(end_col, start_col - 1, -1):
        if is_col_hidden(sheet, col):
            continue
        for row in range(start_row, end_row + 1):
            if is_row_hidden(sheet, row):
                continue
            val = sheet.cell(row=row, column=col).value
            if val is not None and pattern.search(str(val).strip()):
                return col
    return None


def find_last_bordered_table(
    sheet, start_row: int, end_row: int, start_col: int, end_col: int
) -> Optional[Tuple[int, int, int, int]]:
    totalj_col = find_totalj_col(sheet, start_row, end_row, start_col, end_col)
    if totalj_col is None:
        return None

    block_end = totalj_col
    block_start = totalj_col

    for col in range(totalj_col - 1, start_col - 1, -1):
        if is_col_hidden(sheet, col):
            continue
        col_bordered = any(
            cell_has_any_border(sheet.cell(row=r, column=col))
            for r in range(start_row, end_row + 1)
            if not is_row_hidden(sheet, r)
        )
        if col_bordered:
            block_start = col
        else:
            break

    return (block_start, block_end, start_row, end_row)


def get_last_table_bounds(sheet, col_limit: int) -> Optional[Tuple[int, int, int, int]]:
    start_row, end_row, start_col, end_col = get_used_bounds(sheet)
    tables = detect_horizontal_tables(sheet)

    if not tables:
        return None

    last_start_col, last_end_col, t_start_row, t_end_row = tables[-1]

    bordered = find_last_bordered_table(
        sheet, t_start_row, t_end_row, last_start_col, last_end_col
    )

    if bordered:
        b_start_col, b_end_col, b_start_row, b_end_row = bordered
        # Get only visible columns in the bordered block
        vis_cols = get_visible_cols(sheet, b_start_col, b_end_col)
        if col_limit < len(vis_cols):
            # Take only last col_limit visible columns
            vis_cols = vis_cols[-col_limit:]
        return (vis_cols[0], vis_cols[-1], b_start_row, b_end_row)
    else:
        vis_cols = get_visible_cols(sheet, last_start_col, last_end_col)
        if col_limit < len(vis_cols):
            vis_cols = vis_cols[-col_limit:]
        return (vis_cols[0], vis_cols[-1], t_start_row, t_end_row)


# ============================================================
# ========== BOUNDS / VISIBILITY HELPERS =====================
# ============================================================


def get_used_bounds(sheet) -> Tuple[int, int, int, int]:
    """Return (min_row, max_row, min_col, max_col) for the sheet's used range."""
    return sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column


def get_visible_rows(sheet, start_row: int, end_row: int) -> List[int]:
    """Return list of row indices that are not hidden."""
    result = []
    for r in range(start_row, end_row + 1):
        rd = sheet.row_dimensions.get(r)
        if rd and rd.hidden:
            continue
        result.append(r)
    return result


def get_visible_cols(sheet, start_col: int, end_col: int) -> List[int]:
    """Return list of column indices that are not hidden."""
    result = []
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        cd = sheet.column_dimensions.get(col_letter)
        if cd and cd.hidden:
            continue
        result.append(c)
    return result


def _column_has_data(sheet, col: int, start_row: int, end_row: int) -> bool:
    for row in range(start_row, end_row + 1):
        val = sheet.cell(row=row, column=col).value
        if val is not None and str(val).strip() != "":
            return True
    return False


def detect_horizontal_tables(sheet) -> List[Tuple[int, int, int, int]]:
    """
    Detect contiguous column blocks that contain data.
    Returns list of (start_col, end_col, start_row, end_row).
    """
    start_row, end_row, start_col, end_col = get_used_bounds(sheet)
    tables = []
    block_start = None

    for col in range(start_col, end_col + 2):
        has_data = col <= end_col and _column_has_data(sheet, col, start_row, end_row)
        if has_data and block_start is None:
            block_start = col
        elif not has_data and block_start is not None:
            tables.append((block_start, col - 1, start_row, end_row))
            block_start = None

    return tables


def find_last_bordered_table(
    sheet, t_sr: int, t_er: int, t_sc: int, t_ec: int
) -> Optional[Tuple[int, int, int, int]]:
    """
    Within the rectangle (t_sr..t_er, t_sc..t_ec), find the tightest
    bounding box of cells that have any border edge.
    Returns (start_col, end_col, start_row, end_row) or None if no borders found.
    """
    min_r = min_c = None
    max_r = max_c = None

    for r in range(t_sr, t_er + 1):
        for c in range(t_sc, t_ec + 1):
            cell = sheet.cell(row=r, column=c)
            edges = get_border_edges_styled(cell)
            if any(v is not None for v in edges.values()):
                if min_r is None or r < min_r:
                    min_r = r
                if max_r is None or r > max_r:
                    max_r = r
                if min_c is None or c < min_c:
                    min_c = c
                if max_c is None or c > max_c:
                    max_c = c

    if min_r is None:
        return None
    return (min_c, max_c, min_r, max_r)


# ============================================================
# ========== WHATSAPP (GREEN API) ============================
# ============================================================


def get_whatsapp_groups_greenapi(id_instance: str, api_token: str) -> List[dict]:
    """Fetch all WhatsApp groups via Green API. Returns list of {id, name}."""
    url = f"https://api.green-api.com/waInstance{id_instance}/getChats/{api_token}"
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        chats = resp.json()
        return [
            {"id": c["id"], "name": c.get("name") or c["id"]}
            for c in chats
            if isinstance(c.get("id"), str) and c["id"].endswith("@g.us")
        ]
    except Exception as exc:
        logging.error(f"Green API getChats error: {exc}")
        return []


def send_images_to_whatsapp(
    image_paths: List[str],
    chat_id: str,
    id_instance: str,
    api_token: str,
    log_fn,
    delay_seconds: float = 1.5,
) -> Tuple[int, int]:
    """
    Upload and send each image in *image_paths* to *chat_id* via Green API.
    Returns (ok_count, fail_count).
    """
    ok = fail = 0
    upload_url = (
        f"https://api.green-api.com/waInstance{id_instance}"
        f"/sendFileByUpload/{api_token}"
    )

    for img_path in image_paths:
        fname = Path(img_path).name
        try:
            with open(img_path, "rb") as fh:
                resp = requests.post(
                    upload_url,
                    files={"file": (fname, fh, "image/png")},
                    data={"chatId": chat_id, "caption": Path(img_path).stem},
                    timeout=30,
                )
            resp.raise_for_status()
            log_fn(f"✅ Sent: {fname}")
            ok += 1
        except Exception as exc:
            log_fn(f"❌ Failed: {fname} — {exc}")
            fail += 1

        time.sleep(delay_seconds)

    return ok, fail


# ============================================================
# ========== CELL STYLE HELPERS ==============================
# ============================================================


def get_col_width_pixels(sheet, col: int, default_width: float = 8.43) -> float:
    try:
        col_letter = get_column_letter(col)
        cd = sheet.column_dimensions.get(col_letter)
        if cd and cd.width and cd.width > 0:
            return cd.width
    except Exception:
        pass
    return default_width


def get_row_height_pixels(sheet, row: int, default_height: float = 15.0) -> float:
    try:
        rd = sheet.row_dimensions.get(row)
        if rd and rd.height and rd.height > 0:
            return rd.height
    except Exception:
        pass
    return default_height


def get_merged_cell_info(sheet) -> dict:
    """
    Returns a dict mapping every (row, col) in a merged range:
      - top-left cell  → (min_row, min_col, max_row, max_col)
      - slave cells    → None
    """
    merge_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_r = merged_range.min_row
        max_r = merged_range.max_row
        min_c = merged_range.min_col
        max_c = merged_range.max_col
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    merge_map[(r, c)] = (min_r, min_c, max_r, max_c)
                else:
                    merge_map[(r, c)] = None  # slave
    return merge_map


def get_cell_alignment(cell):
    """Returns (horizontal, vertical, wrap_text)."""
    h_align, v_align, wrap = "left", "center", False
    try:
        aln = cell.alignment
        if aln:
            h = aln.horizontal or "left"
            h_align = h if h != "general" else "left"
            v_align = aln.vertical or "center"
            wrap = bool(aln.wrap_text)
    except Exception:
        pass
    return h_align, v_align, wrap


def get_cell_font_size(cell, default: float = 10.0) -> float:
    try:
        if cell.font and cell.font.size:
            return float(cell.font.size)
    except Exception:
        pass
    return default


def get_cell_font_italic(cell) -> bool:
    try:
        return bool(cell.font.italic) if cell.font else False
    except Exception:
        return False


def get_cell_bg_color(cell) -> str:
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            fg = fill.fgColor
            if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
                return f"#{fg.rgb[2:]}"
            elif fg.type == "theme":
                return "#FFFFFF"
    except Exception:
        pass
    return "#FFFFFF"


def get_cell_font_color(cell) -> str:
    try:
        font = cell.font
        if font and font.color:
            fc = font.color
            if fc.type == "rgb" and fc.rgb and fc.rgb != "00000000":
                return f"#{fc.rgb[2:]}"
    except Exception:
        pass
    return "#000000"


def get_cell_font_bold(cell) -> bool:
    try:
        return cell.font.bold if cell.font else False
    except Exception:
        return False


def get_border_edges_styled(cell) -> dict:
    """Returns per-side border info dict with linewidth and color, or None."""
    THICK_STYLES = {
        "medium",
        "thick",
        "double",
        "mediumDashed",
        "mediumDashDot",
        "mediumDashDotDot",
        "slantDashDot",
    }
    edges = {"left": None, "right": None, "top": None, "bottom": None}
    try:
        b = cell.border
        for side in edges:
            s = getattr(b, side, None)
            if s and s.border_style and s.border_style != "none":
                lw = 1.5 if s.border_style in THICK_STYLES else 0.7
                clr = "#000000"
                try:
                    if s.color and s.color.rgb and s.color.rgb != "00000000":
                        clr = f"#{s.color.rgb[2:]}"
                except Exception:
                    pass
                edges[side] = {"lw": lw, "color": clr}
    except Exception:
        pass
    return edges


# ============================================================
# ========== UNIT CONVERSION =================================
# ============================================================


def excel_col_width_to_inches(excel_width: float) -> float:
    """1 Excel column-width unit ≈ 0.14 inches (default font, 96 dpi)."""
    return excel_width * 0.14


def excel_row_height_to_inches(pt_height: float) -> float:
    """Excel row height is in points; 1 pt = 1/72 inch."""
    return pt_height / 72.0


# ============================================================
# ========== IMAGE EXPORT ====================================
# ============================================================


def export_range_as_image(
    sheet,
    rng_bounds: Tuple[int, int, int, int],
    output_path: str,
    visible_rows: Optional[List[int]] = None,
    visible_cols: Optional[List[int]] = None,
) -> None:
    """
    Renders the visible portion of *rng_bounds* from *sheet* into a PNG.

    rng_bounds : (start_row, end_row, start_col, end_col)  — 1-indexed
    visible_rows / visible_cols : pre-filtered lists; computed if None.

    Handles:
      • merged cells (correct span geometry)
      • real column widths & row heights
      • per-cell font (size, bold, italic, colour)
      • per-cell alignment (H/V) and text-wrap
      • background fill colours
      • borders (thin / thick, custom colours)
    """
    import textwrap

    start_row, end_row, start_col, end_col = rng_bounds

    # ── Visible row / col lists ──────────────────────────────────────
    if visible_rows is None:
        visible_rows = get_visible_rows(sheet, start_row, end_row)
    if visible_cols is None:
        visible_cols = get_visible_cols(sheet, start_col, end_col)

    visible_rows = [r for r in visible_rows if start_row <= r <= end_row]
    visible_cols = [c for c in visible_cols if start_col <= c <= end_col]

    n_rows = len(visible_rows)
    n_cols = len(visible_cols)
    if n_rows == 0 or n_cols == 0:
        return

    # ── Physical sizes (inches) ──────────────────────────────────────
    col_widths_in = [
        excel_col_width_to_inches(get_col_width_pixels(sheet, c)) for c in visible_cols
    ]
    row_heights_in = [
        excel_row_height_to_inches(get_row_height_pixels(sheet, r))
        for r in visible_rows
    ]

    fig_width = max(sum(col_widths_in), 1.0)
    fig_height = max(sum(row_heights_in), 0.5)

    # ── Coordinate grids ────────────────────────────────────────────
    # col_x[i]  = left  edge of visible_cols[i]  (in inches)
    # col_x[n]  = right edge of last column
    col_x = [0.0]
    for w in col_widths_in:
        col_x.append(col_x[-1] + w)

    # row_y_top[i]    = top    edge of visible_rows[i]  (matplotlib y, 0=bottom)
    # row_y_bottom[i] = bottom edge of visible_rows[i]
    row_y_top = [0.0] * n_rows
    row_y_bottom = [0.0] * n_rows
    y = fig_height
    for i, h in enumerate(row_heights_in):
        row_y_top[i] = y
        row_y_bottom[i] = y - h
        y -= h

    # ── Merge map & index maps ───────────────────────────────────────
    merge_map = get_merged_cell_info(sheet)
    row_idx_map = {r: i for i, r in enumerate(visible_rows)}
    col_idx_map = {c: i for i, c in enumerate(visible_cols)}

    # ── Figure ───────────────────────────────────────────────────────
    DPI = 150
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=DPI)
    ax.set_xlim(0, fig_width)
    ax.set_ylim(0, fig_height)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    drawn = set()  # (row, col) cells already rendered

    for r_idx, r in enumerate(visible_rows):
        for c_idx, c in enumerate(visible_cols):
            if (r, c) in drawn:
                continue

            merge_info = merge_map.get((r, c), "NOT_MERGED")

            # ── Slave cell → skip ────────────────────────────────
            if merge_info is None:  # slave
                drawn.add((r, c))
                continue

            cell = sheet.cell(row=r, column=c)

            # ── Compute rect geometry ────────────────────────────
            if merge_info != "NOT_MERGED":  # top-left of a merge
                min_r, min_c, max_r, max_c = merge_info

                # Collect visible span indices
                span_r_idxs = [
                    row_idx_map[rr]
                    for rr in range(min_r, max_r + 1)
                    if rr in row_idx_map
                ]
                span_c_idxs = [
                    col_idx_map[cc]
                    for cc in range(min_c, max_c + 1)
                    if cc in col_idx_map
                ]

                if not span_r_idxs or not span_c_idxs:
                    drawn.add((r, c))
                    continue

                x_left = col_x[span_c_idxs[0]]
                x_right = col_x[span_c_idxs[-1] + 1]
                y_top = row_y_top[span_r_idxs[0]]
                y_bottom = row_y_bottom[span_r_idxs[-1]]

                # Mark every cell in the merge as drawn
                for mr in range(min_r, max_r + 1):
                    for mc in range(min_c, max_c + 1):
                        drawn.add((mr, mc))
            else:  # normal (non-merged) cell
                x_left = col_x[c_idx]
                x_right = col_x[c_idx + 1]
                y_top = row_y_top[r_idx]
                y_bottom = row_y_bottom[r_idx]
                drawn.add((r, c))

            cell_w = x_right - x_left
            cell_h = y_top - y_bottom

            # ── Background fill ──────────────────────────────────
            bg = get_cell_bg_color(cell)
            ax.add_patch(
                plt.Rectangle(
                    (x_left, y_bottom),
                    cell_w,
                    cell_h,
                    facecolor=bg,
                    edgecolor="none",
                    linewidth=0,
                    zorder=1,
                )
            )

            # ── Borders ──────────────────────────────────────────
            edges = get_border_edges_styled(cell)
            border_segments = [
                ("top", [x_left, x_right], [y_top, y_top]),
                ("bottom", [x_left, x_right], [y_bottom, y_bottom]),
                ("left", [x_left, x_left], [y_bottom, y_top]),
                ("right", [x_right, x_right], [y_bottom, y_top]),
            ]
            for side, xs, ys in border_segments:
                info = edges[side]
                if info:
                    ax.plot(xs, ys, color=info["color"], linewidth=info["lw"], zorder=3)

            # ── Text ─────────────────────────────────────────────
            val = cell.value
            if val is None or str(val).strip() == "":
                continue

            text_str = str(val)
            fg = get_cell_font_color(cell)
            bold = get_cell_font_bold(cell)
            italic = get_cell_font_italic(cell)
            font_pt = get_cell_font_size(cell, default=10.0)
            font_size = font_pt * 0.82  # slight scale for tight layout

            h_align, v_align, wrap_text = get_cell_alignment(cell)

            HA_MAP = {
                "left": "left",
                "center": "center",
                "right": "right",
                "general": "left",
                "fill": "left",
                "justify": "left",
                "distributed": "center",
            }
            VA_MAP = {
                "top": "top",
                "center": "center",
                "bottom": "bottom",
                "justify": "center",
                "distributed": "center",
            }

            ha = HA_MAP.get(h_align, "left")
            va = VA_MAP.get(v_align, "center")

            PAD = 0.04  # inches padding inside cell
            tx = (
                x_left + PAD
                if ha == "left"
                else x_right - PAD
                if ha == "right"
                else (x_left + x_right) / 2
            )
            ty = (
                y_top - PAD
                if va == "top"
                else y_bottom + PAD
                if va == "bottom"
                else (y_top + y_bottom) / 2
            )

            # ── Text wrapping ────────────────────────────────────
            if wrap_text and cell_w > 0:
                char_w_in = font_pt * 0.55 / 72.0  # avg char width
                chars_per_line = max(1, int((cell_w - 2 * PAD) / char_w_in))
                display_text = textwrap.fill(text_str, width=chars_per_line)
            else:
                display_text = text_str

            ax.text(
                tx,
                ty,
                display_text,
                ha=ha,
                va=va,
                fontsize=font_size,
                color=fg,
                fontweight="bold" if bold else "normal",
                fontstyle="italic" if italic else "normal",
                clip_on=True,
                zorder=4,
                multialignment=ha,
                linespacing=1.2,
            )

    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
    fig.savefig(output_path, dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)


# ============================================================
# ========== CORE RECAP FUNCTION =============================
# ============================================================


def get_recap(
    sheet,
    workbook_name: str,
    output_folder: str,
    last_table_tail_cols: int,
    log_fn,  # callable(str) – e.g. log_area.write
) -> List[str]:
    """
    Detects all horizontal tables in *sheet*, exports:
      • one PNG per non-last table  (full visible extent)
      • two PNGs for the last table:
          1. full bordered block
          2. tail (last *last_table_tail_cols* visible columns of the block)

    Returns a list of absolute paths to the generated PNG files.
    """
    generated: List[str] = []

    tables = detect_horizontal_tables(sheet)
    if not tables:
        log_fn(f"⚠️  No tables found in '{workbook_name}'")
        return generated

    # Pre-compute visible rows/cols for the whole used range once
    s_row, e_row, s_col, e_col = get_used_bounds(sheet)
    all_vis_rows = get_visible_rows(sheet, s_row, e_row)
    all_vis_cols = get_visible_cols(sheet, s_col, e_col)

    sheet_title = sheet.title

    # ── Non-last tables: full picture ───────────────────────────────
    for i, (t_sc, t_ec, t_sr, t_er) in enumerate(tables[:-1], start=1):
        rng = (t_sr, t_er, t_sc, t_ec)
        fname = f"{workbook_name}__{sheet_title}__table{i}.png"
        fpath = os.path.join(output_folder, fname)

        export_range_as_image(
            sheet,
            rng,
            fpath,
            visible_rows=all_vis_rows,
            visible_cols=all_vis_cols,
        )
        generated.append(fpath)
        log_fn(f"✅ Saved: {fname}")

    # ── Last table ───────────────────────────────────────────────────
    t_sc, t_ec, t_sr, t_er = tables[-1]
    table_num = len(tables)

    bordered = find_last_bordered_table(sheet, t_sr, t_er, t_sc, t_ec)
    if bordered:
        b_sc, b_ec, b_sr, b_er = bordered
    else:
        b_sc, b_ec, b_sr, b_er = t_sc, t_ec, t_sr, t_er

    # Visible cols inside this block
    block_vis_cols = [c for c in all_vis_cols if b_sc <= c <= b_ec]

    # Picture 1 – full bordered block
    rng_full = (b_sr, b_er, b_sc, b_ec)
    fname_full = f"{workbook_name}__{sheet_title}__table{table_num}_full.png"
    fpath_full = os.path.join(output_folder, fname_full)

    export_range_as_image(
        sheet,
        rng_full,
        fpath_full,
        visible_rows=all_vis_rows,
        visible_cols=block_vis_cols,
    )
    generated.append(fpath_full)
    log_fn(f"✅ Saved (last – full):  {fname_full}")

    # Picture 2 – tail (last N visible cols of the block)
    tail_count = min(last_table_tail_cols, len(block_vis_cols))
    tail_cols = block_vis_cols[-tail_count:]

    if tail_cols:
        rng_tail = (b_sr, b_er, tail_cols[0], tail_cols[-1])
        fname_tail = (
            f"{workbook_name}__{sheet_title}__table{table_num}_tail{tail_count}.png"
        )
        fpath_tail = os.path.join(output_folder, fname_tail)

        export_range_as_image(
            sheet,
            rng_tail,
            fpath_tail,
            visible_rows=all_vis_rows,
            visible_cols=tail_cols,
        )
        generated.append(fpath_tail)
        log_fn(f"✅ Saved (last – tail {tail_count} cols): {fname_tail}")

    return generated