import os
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from assets.constants.constants import (
    PATH_DEBRQ,
    COMMODITY_TYPES,
    COL_CLIENT,
    COL_TYPE,
    COL_QUANTITE,
    COL_BL,
    GOODS__TYPES,
    COL_DATE,
    COL_MODELE, COL_CHASSIS_SERIAL, COL_PRODUIT,
    COL_DESIGNATION,
    KEYWORD_RULES
)
from tools.tools import group_sourcefile_by_client,apply_summary_conditional_formatting,get_manual_color

# to DEFINE THE TITLE OF SHIP
def style_header_cell(ws, text, cell_range="C1:J1", bg_color="D3D3D3", font_color="0000FF"):
    # Define styles
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    font = Font(name="Times New Roman",bold=True, size=18, color=font_color)
    align = Alignment(horizontal="center", vertical="center")
    
    # Apply logic
    ws.merge_cells(cell_range)
    top_left_cell = cell_range.split(":")[0]
    
    
    ws[top_left_cell].value = text
    ws[top_left_cell].fill = fill
    ws[top_left_cell].font = font
    ws[top_left_cell].alignment = align



def create_product_table(ws, product_name, product_data, start_col, is_others=False):
    # --- Color Selection ---
    raw_color = get_manual_color(product_name) if  is_others is False else None

    # --- Styles ---
    header_fill = PatternFill(
        start_color=raw_color,
        end_color=raw_color,
        fill_type="solid"
    ) if raw_color else PatternFill(fill_type=None)

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Column Setup (use constants) ---
    client_col = COL_CLIENT.upper()
    prod_col = COL_TYPE.upper()

    clients = product_data[client_col].unique().tolist()

    col_mapping = {}
    for i, client in enumerate(clients):
        col_letter = get_column_letter(start_col + 2 + i)
        col_mapping[client] = col_letter
        ws.column_dimensions[col_letter].width = 22

    ws.column_dimensions[get_column_letter(start_col)].width = 15
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 12

    num_extra_cols = 3
    start_col_idx = start_col + 2
    last_col_idx = start_col_idx + len(clients) + num_extra_cols - 1
    last_col_letter = get_column_letter(last_col_idx)

    # --- Row 4: Product Title Header ---
    ws.merge_cells(f"{get_column_letter(start_col_idx)}4:{last_col_letter}4")
    title_cell = ws[f"{get_column_letter(start_col_idx)}4"]
    title_cell.value = product_name
    title_cell.font = title_font
    title_cell.alignment = center
    title_cell.fill = header_fill
    for col_i in range(start_col_idx, last_col_idx + 1):
        ws.cell(row=4, column=col_i).border = border

    # --- Row 5: Client Names ---
    for c_idx in range(start_col_idx, last_col_idx + 1):
        cell = ws.cell(row=5, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for client, col in col_mapping.items():
        ws[f"{col}5"].value = str(client)

    # --- Row 6: Alignment logic ---
    if is_others:
        for client, col in col_mapping.items():
            try:
                prod_val = product_data[product_data[client_col] == client][prod_col].iloc[0]
            except Exception:
                prod_val = "-"
            c = ws[f"{col}6"]
            c.value = str(prod_val)
            c.font = Font(size=10, italic=True)
            c.alignment = center
            c.border = border
    else:
        for c_idx in range(start_col_idx, last_col_idx + 1):
            ws.cell(row=6, column=c_idx).border = border

    # --- Row 7: BL Numbers / Headers ---
    bl_row = 7
    for col_idx in range(start_col, last_col_idx + 1):
        cell = ws.cell(row=bl_row, column=col_idx)
        cell.border = border
        cell.alignment = center
        if col_idx >= start_col_idx:
            cell.fill = header_fill
            cell.font = header_font

    ws[f"{get_column_letter(start_col)}{bl_row}"].value = "DATE"
    ws[f"{get_column_letter(start_col + 1)}{bl_row}"].value = "SHIFT"

    # Use constants for BL and quantity columns (matching normalized DF columns)
    bl_col_name = COL_BL.upper()
    qty_col_name = COL_QUANTITE.upper()

    for client, col in col_mapping.items():
        try:
            bl_num = product_data[product_data[client_col] == client][bl_col_name].iloc[0]
        except Exception:
            bl_num = "-"
        ws[f"{col}{bl_row}"].value = str(bl_num)

    extra_headers = ["INC", "TOTAL", "TOTAL/J"]
    extra_cols = []
    for i, h in enumerate(extra_headers):
        col_idx = start_col + 2 + len(clients) + i
        col_letter = get_column_letter(col_idx)
        extra_cols.append(col_letter)
        ws[f"{col_letter}{bl_row}"].value = h

    # --- Data Rows (Starting at Row 8 for 15 Days) ---
    

    
    # base_date = product_data[COL_DATE].iloc[0] if not product_data[COL_DATE].empty else datetime.now()
    # Ensure base_date is a datetime object, not a string
    base_date = pd.to_datetime(product_data[COL_DATE].iloc[0]) if not product_data[COL_DATE].empty else datetime.now()

    shifts = ["MATIN", "SOIR", "NUIT", "NUIT -2-"]
    data_start_row = 8
    curr_data_row = data_start_row

    for day_offset in range(15):
        # ADD THIS: Determine if row should be hidden
        should_hide = day_offset > 0
        d_str = (base_date + timedelta(days=day_offset)).strftime("%Y-%m-%d")
        day_start_row = curr_data_row
        for shift in shifts:
            # ADD THIS: Hide the row dimension
            if should_hide:
                ws.row_dimensions[curr_data_row].hidden = True

            ws[f"{get_column_letter(start_col)}{curr_data_row}"].value = d_str
            ws[f"{get_column_letter(start_col + 1)}{curr_data_row}"].value = shift

            for client, col in col_mapping.items():
                ws[f"{col}{curr_data_row}"].value = 0

            inc_col = extra_cols[0]
            ws[f"{inc_col}{curr_data_row}"].value = 0

            total_col = extra_cols[1]
            first_client_col = get_column_letter(start_col + 2)
            ws[f"{total_col}{curr_data_row}"].value = (
                f"=SUM({first_client_col}{curr_data_row}:{inc_col}{curr_data_row})"
            )

            for c_idx in range(start_col, last_col_idx + 1):
                cell = ws.cell(row=curr_data_row, column=c_idx)
                cell.border = border
                cell.alignment = center

            curr_data_row += 1

        ws.merge_cells(
            f"{get_column_letter(start_col)}{day_start_row}:"
            f"{get_column_letter(start_col)}{curr_data_row - 1}"
        )
        totalj_col = extra_cols[2]
        ws.merge_cells(
            f"{totalj_col}{day_start_row}:{totalj_col}{curr_data_row - 1}"
        )
        ws[f"{totalj_col}{day_start_row}"].value = (
            f"=SUM({extra_cols[1]}{day_start_row}:{extra_cols[1]}{day_start_row+3})"
        )
        ws[f"{totalj_col}{day_start_row}"].alignment = center
        ws[f"{totalj_col}{day_start_row}"].font = Font(bold=True)
        ws[f"{totalj_col}{day_start_row}"].border = border

    summary_start_row = curr_data_row
    labels = ["TOTAL DECHARGER", "QUANTITE MANIFEST", "RESTE A BORD"]
    summary_rows = []
    for i, label in enumerate(labels):
        r = summary_start_row + i
        col1 = start_col
        col2 = start_col + 1
        col1_letter = get_column_letter(col1)
        col2_letter = get_column_letter(col2)

        # Merge label cell with the next right cell
        ws.merge_cells(f"{col1_letter}{r}:{col2_letter}{r}")

        label_cell = ws[f"{col1_letter}{r}"]
        label_cell.value = label
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = border

        # Border on the second cell (even though merged, needed for right border)
        ws[f"{col2_letter}{r}"].border = border

        summary_rows.append(r)


    # for client, col in list(col_mapping.items()) + [(None, extra_cols[0])]:
    for client, col in col_mapping.items():
        # TOTAL DECHARGER
        ws[f"{col}{summary_rows[0]}"].value = (
            f"=SUM({col}{data_start_row}:{col}{curr_data_row-1})"
        )

        target_data = product_data[product_data[client_col] == client]
        q_manifest = pd.to_numeric(
            target_data[qty_col_name], errors="coerce"
        ).sum()

        # QUANTITE MANIFEST
        ws[f"{col}{summary_rows[1]}"].value = q_manifest
        # RESTE A BORD
        ws[f"{col}{summary_rows[2]}"].value = (
            f"={col}{summary_rows[1]}-{col}{summary_rows[0]}"
        )

        for r in summary_rows:
            ws[f"{col}{r}"].font = Font(bold=True)
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].alignment = center

    # handling the INC COLS
    # INC column summary (separate handling)
    inc_col = extra_cols[0]
    # TOTAL DECHARGER for INC
    ws[f"{inc_col}{summary_rows[0]}"].value = (
        f"=SUM({inc_col}{data_start_row}:{inc_col}{curr_data_row-1})"
    )
    
    # QUANTITE MANIFEST for INC (total of all product data)
    ws[f"{inc_col}{summary_rows[1]}"].value = 0
    
    # RESTE A BORD for INC
    ws[f"{inc_col}{summary_rows[2]}"].value = (
        f"={inc_col}{summary_rows[1]}-{inc_col}{summary_rows[0]}"
    )
    for r in summary_rows:
        ws[f"{inc_col}{r}"].font = Font(bold=True)
        ws[f"{inc_col}{r}"].border = border
        ws[f"{inc_col}{r}"].alignment = center


    # Global totals columns (TOTAL / TOTAL/J)
    for r in summary_rows:
        ws[f"{extra_cols[1]}{r}"].value = (f"=SUM({get_column_letter(start_col+2)}{r}:{extra_cols[0]}{r})")
        ws[f"{extra_cols[1]}{r}"].font = Font(bold=True)
        ws[f"{extra_cols[1]}{r}"].border = border
        ws[f"{extra_cols[1]}{r}"].alignment = center

    apply_summary_conditional_formatting(ws, summary_rows, start_col, clients)

    
    # return last_col_idx
    return last_col_idx, summary_rows, extra_cols[1] 

def gen_table_deb(filepath=None):

    if not filepath:
        return False

    base_name = os.path.basename(filepath)
    file_name_only = os.path.splitext(base_name)[0].upper()

    list_bl = pd.read_excel(filepath, sheet_name=0, engine="openpyxl")

    source_df = group_sourcefile_by_client(filepath, skip_units_packages=True, bl_aggregated=False)
    
    source_df.columns = source_df.columns.str.strip().str.upper()
    st.dataframe(source_df)

    list_bl.columns = list_bl.columns.str.strip().str.upper()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{file_name_only}"

    ship_name_placeholder = f"SHIP NAME:      {file_name_only}"

    list_bl_sheet_name = f"LIST_BL_{file_name_only}"
    ws_bl = wb.create_sheet(title=list_bl_sheet_name)
   
    # ─── Styles ───────────────────────────────────────────────────────────────
    thick_side = Side(border_style="thick", color="000000")
    thin_side  = Side(border_style="thin",  color="000000")

    header_border = Border(top=thick_side, bottom=thick_side,
                           left=thick_side, right=thick_side)
    data_border   = Border(top=thin_side,  bottom=thin_side,
                           left=thin_side,  right=thin_side)

    center_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Times New Roman 12 pt – base font for ws_bl
    base_font        = Font(name="Times New Roman", size=12)
    base_font_bold   = Font(name="Times New Roman", size=12, bold=True)

    # Purple font for the CLIENT column
    purple_font      = Font(name="Times New Roman", size=12,
                            bold=True, color="7030A0")   # standard purple

    ROW_HEIGHT = 15
    COL_WIDTH  = 20

    # ─── Column-index helpers ─────────────────────────────────────────────────
    cols = list(list_bl.columns)

    type_col_idx    = list_bl.columns.get_loc(COL_TYPE)          # 0-based

    # Columns that receive row_fill coloring (0-based indices)
    colored_col_indices = set()
    for col_const in (COL_TYPE, COL_MODELE, COL_CHASSIS_SERIAL, COL_PRODUIT):
        if col_const in list_bl.columns:
            colored_col_indices.add(list_bl.columns.get_loc(col_const))

    # Client column index (0-based), if it exists
    client_col_idx = list_bl.columns.get_loc(COL_CLIENT) if COL_CLIENT in list_bl.columns else None

    # ─── Write Headers ────────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(list_bl.columns, start=1):
        cell = ws_bl.cell(row=1, column=c_idx)
        cell.value     = col_name
        cell.font      = base_font_bold          # Times New Roman 12 bold
        cell.border    = header_border
        cell.alignment = center_alignment
        ws_bl.column_dimensions[cell.column_letter].width = COL_WIDTH

    # ─── Write Data with Styling ──────────────────────────────────────────────
    for r_idx, row in enumerate(list_bl.itertuples(index=False), start=2):
        ws_bl.row_dimensions[r_idx].height = ROW_HEIGHT

        # Determine row fill colour from COL_TYPE value
        hex_color = get_manual_color(row[type_col_idx])
        row_fill  = (
            PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            if hex_color else None
        )

        for c_idx, value in enumerate(row, start=1):
            zero_idx = c_idx - 1          # 0-based column index
            cell = ws_bl.cell(row=r_idx, column=c_idx)

            # ── Value ──
            # Force CLIENT column to uppercase string
            if zero_idx == client_col_idx and value is not None:
                cell.value = str(value).upper()
            else:
                cell.value = value

            # ── Alignment & Border (all cells) ──
            cell.alignment = center_alignment
            cell.border    = data_border

            # ── Font ──
            if zero_idx == client_col_idx:
                cell.font = purple_font          # purple + Times New Roman 12
            else:
                cell.font = base_font            # Times New Roman 12

            # ── Fill – only on the four designated columns ──
            if row_fill and zero_idx in colored_col_indices:
                cell.fill = row_fill

    # ─── Rest of the function (ws / deb sheet) is unchanged ──────────────────
    style_header_cell(ws, ship_name_placeholder)

    specific_keywords = GOODS__TYPES

    start_col = 1
    all_matched_indices = pd.Index([])
    all_summary_info = []

    for keyword in specific_keywords:
        mask   = source_df[COL_TYPE].astype(str).str.contains(keyword, case=False, na=False)
        p_data = source_df[mask]

        if not p_data.empty:
            all_matched_indices = all_matched_indices.union(p_data.index)
            last_col_idx, summary_rows, total_col = create_product_table(
                ws, keyword.upper(), p_data, start_col, is_others=False
            )
            all_summary_info.append((summary_rows, total_col))
            start_col = last_col_idx + 3

    others_data = source_df.drop(all_matched_indices)
    if not others_data.empty:
        last_col_idx, summary_rows, total_col = create_product_table(
            ws, "UNITS + PACKAGES", others_data, start_col, is_others=True
        )
        all_summary_info.append((summary_rows, total_col))

    if all_summary_info:
        max_row = max(max(rows) for rows, _ in all_summary_info)
        global_summary_start_row = max_row + 3

        border      = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"),  bottom=Side(style="thin"))
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font   = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        label_fill  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        header_row = global_summary_start_row
        ws.merge_cells(f"A{header_row}:B{header_row}")
        ws[f"A{header_row}"].value = "GLOBAL SUMMARY"
        for col_idx in range(1, 3):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill      = header_fill
            cell.font      = Font(bold=True, size=13)
            cell.border    = border
            cell.alignment = center

        row_total    = header_row + 1
        row_manifest = header_row + 2
        row_reste    = header_row + 3

        ws[f"A{row_total}"].value = "TOTAL DECHARGER"
        total_decharger_refs = ",".join(f"{total_col}{rows[0]}" for rows, total_col in all_summary_info)
        ws[f"B{row_total}"].value = (f"={total_decharger_refs}" if len(all_summary_info) == 1
                                     else f"=SUM({total_decharger_refs})")

        ws[f"A{row_manifest}"].value = "QUANTITE MANIFEST"
        quantite_refs = ",".join(f"{total_col}{rows[1]}" for rows, total_col in all_summary_info)
        ws[f"B{row_manifest}"].value = (f"={quantite_refs}" if len(all_summary_info) == 1
                                        else f"=SUM({quantite_refs})")

        ws[f"A{row_reste}"].value = "RESTE A BORD"
        reste_refs = ",".join(f"{total_col}{rows[2]}" for rows, total_col in all_summary_info)
        ws[f"B{row_reste}"].value = (f"={reste_refs}" if len(all_summary_info) == 1
                                     else f"=SUM({reste_refs})")

        for r in (row_total, row_manifest, row_reste):
            label_cell           = ws.cell(row=r, column=1)
            label_cell.fill      = label_fill
            label_cell.font      = bold_font
            label_cell.border    = border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=r, column=2)
            value_cell.font      = bold_font
            value_cell.border    = border
            value_cell.alignment = center

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 22

    output_xlsx = f"{PATH_DEBRQ}/{file_name_only}.xlsx"
    wb.save(output_xlsx)

    return output_xlsx