# ============================================================
# PAGE: Generate_Sheets
# ============================================================

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os
import re
import io
import threading
import streamlit as st
from difflib import SequenceMatcher
from typing import Optional

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TODAY           = date.today()
TODAY_STR       = TODAY.strftime("%d-%m-%Y")
RST_SHEET_NAME  = "RST-10-10-2026"
SOURCE_FILE     = "manifest_source.xlsx"
OUTPUT_FILE     = "output_omar.xlsx"

# ---------------------------------------------------------------------------
# ══════════════════════════════════════════════════════════════════════════
#  FUZZY HEADER RESOLVER
#  Strategy:
#   1. Exact match (after strip)
#   2. Case-insensitive match
#   3. Normalized match  (collapse spaces, remove punctuation)
#   4. Token-subset match (all tokens of canonical are in column tokens)
#   5. Character-ratio fuzzy match via SequenceMatcher (≥ 0.82 threshold)
#  Each canonical name maps to a list of aliases tried in order.
# ══════════════════════════════════════════════════════════════════════════
# ---------------------------------------------------------------------------

# Canonical name → list of known aliases / alternate spellings
COLUMN_ALIASES: dict[str, list[str]] = {
    "escale_numb":        ["escale_numb", "escale numb", "n escale", "n° escale",
                           "numero escale", "numéro escale", "escale"],
    "navire_name":        ["navire_name", "navire name", "navire", "vessel", "ship name"],
    "CONSIGNATAIRE":      ["consignataire", "consignatair", "consign"],
    "DATE":               ["date", "dat"],
    "DATE D'ENTREE":      ["date d'entree", "date entree", "date d entree",
                           "date d'entrée", "date entrée", "date arrivee",
                           "date arrivée", "date d'arrivee"],
    "EMBARQ/DEBARQ":      ["embarq/debarq", "embarq debarq", "embarq / debarq",
                           "emb/deb", "embarquement debarquement", "emb deb"],
    "PRODUITS":           ["produits", "produit", "product", "marchandise type"],
    "Détails PRODUITS":   ["détails produits", "details produits", "detail produits",
                           "détail produits", "détails produit", "details produit",
                           "produits details", "produit detail"],
    "nombre colis":       ["nombre colis", "nbre colis", "nb colis", "nombre de colis",
                           "nbr colis", "nombre col", "colis", "qte", "quantite",
                           "quantité"],
    "Poids brute":        ["poids brute", "poids brut", "poid brute", "poid brut",
                           "tonnage", "weight", "gross weight", "poids"],
    "Client":             ["client", "clients", "clien", "customer", "nom client"],
    "BL":                 ["bl", "b/l", "b l", "bill of lading", "bl numero",
                           "numéro bl", "numero bl", "n° bl"],
    "Article":            ["article", "articles", "art", "crn", "crn/art", "crn art"],
    "Marchandise":        ["marchandise", "marchandises", "marchandise d",
                           "marchandi", "goods", "designation marchandise"],
    "Marchandise.1":      ["marchandise.1", "marchandise 1", "marchandise h",
                           "marchandises.1"],
    "Némuro de chassis":  ["némuro de chassis", "numero de chassis", "numéro de chassis",
                           "n° chassis", "n chassis", "chassis", "chassis number",
                           "vin", "numro de chassis", "nemuro de chassis",
                           "numero chassis", "numéro chassis", "chassis no"],
    "Modèle":             ["modèle", "modele", "model", "mod", "modèl"],
    "SURF":               ["surf", "surface", "superficie", "surface m2", "surf m2"],
    "MRN":                ["mrn", "n° gros", "n gros", "gros", "mrn / n° gros",
                           "mrn/n° gros"],
}

# Reverse lookup built once: normalised_alias → canonical
def _build_reverse_map() -> dict[str, str]:
    rev: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            rev[_norm(alias)] = canonical
    return rev

def _norm(s: str) -> str:
    """Lowercase, collapse whitespace, remove common punctuation."""
    s = s.lower().strip()
    s = re.sub(r"[°'\"./\\]", " ", s)   # replace punctuation with space
    s = re.sub(r"\s+", " ", s)           # collapse spaces
    return s.strip()

def _tokenize(s: str) -> set[str]:
    return set(re.split(r"\s+", _norm(s)))

_REVERSE_MAP: dict[str, str] = {}   # filled lazily on first use


def resolve_column(col_name: str) -> Optional[str]:
    """
    Given an actual spreadsheet column name, return the canonical key
    (from COLUMN_ALIASES) or None if no match.
    """
    global _REVERSE_MAP
    if not _REVERSE_MAP:
        _REVERSE_MAP = _build_reverse_map()

    # 1. Exact match
    if col_name in COLUMN_ALIASES:
        return col_name

    normed = _norm(col_name)

    # 2. Direct normalised lookup
    if normed in _REVERSE_MAP:
        return _REVERSE_MAP[normed]

    # 3. Token-subset match:
    #    canonical tokens ⊆ actual tokens  (handles extra words in header)
    col_tokens = _tokenize(col_name)
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_tokens = _tokenize(alias)
            if alias_tokens and alias_tokens.issubset(col_tokens):
                return canonical

    # 4. Fuzzy ratio match (SequenceMatcher ≥ 0.82)
    best_ratio   = 0.0
    best_canon   = None
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            ratio = SequenceMatcher(None, normed, _norm(alias)).ratio()
            if ratio > best_ratio:
                best_ratio  = ratio
                best_canon  = canonical
    if best_ratio >= 0.82:
        return best_canon

    return None   # no match found


def build_column_map(df_columns) -> dict[str, str]:
    """
    Returns  {canonical_name: actual_df_column_name}
    for every column in the DataFrame that could be resolved.
    Logs unresolved columns for debugging.
    """
    mapping: dict[str, str] = {}          # canonical → actual col
    unresolved: list[str]   = []

    for col in df_columns:
        canon = resolve_column(str(col))
        if canon:
            # First match wins (handles duplicate-like columns)
            if canon not in mapping:
                mapping[canon] = str(col)
        else:
            unresolved.append(str(col))

    if unresolved:
        # Surface in Streamlit sidebar for debugging — non-blocking
        try:
            with st.sidebar.expander("⚠️ Unresolved columns", expanded=False):
                st.write(unresolved)
        except Exception:
            pass

    return mapping


# ---------------------------------------------------------------------------
# Safe accessor that uses the dynamic column map
# ---------------------------------------------------------------------------
def safe_val(row, canonical: str, col_map: dict[str, str]):
    """Fetch value using canonical name resolved through col_map."""
    actual = col_map.get(canonical)
    if actual is None or actual not in row.index:
        return ""
    val = row[actual]
    return "" if pd.isna(val) else val


# ---------------------------------------------------------------------------
# Multi-sheet loader: reads ALL sheets from manifest_source and stacks them
# ---------------------------------------------------------------------------
def load_all_sheets(filepath: str) -> tuple[pd.DataFrame, list[str], dict]:
    """
    Opens every sheet in `filepath`, normalises headers via the fuzzy resolver,
    renames columns to their canonical names, and concatenates into one DataFrame.

    Returns:
        combined_df  – single DataFrame with canonical column names
        sheet_names  – list of sheet names found
        per_sheet    – dict  {sheet_name: row_count}
    """
    xl      = pd.ExcelFile(filepath)
    sheets  = xl.sheet_names
    frames  = []
    per_sheet: dict[str, int] = {}

    for sheet in sheets:
        try:
            df = xl.parse(sheet, header=0, dtype=str)
        except Exception as e:
            st.warning(f"⚠️ Could not read sheet `{sheet}`: {e}")
            continue

        if df.empty:
            per_sheet[sheet] = 0
            continue

        # Strip whitespace from column names
        df.columns = [str(c).strip() for c in df.columns]

        # Build column map for THIS sheet
        col_map = build_column_map(df.columns)

        # Rename columns to canonical names
        rename_dict = {actual: canon for canon, actual in col_map.items()}
        df = df.rename(columns=rename_dict)

        # Tag with source sheet so downstream code can use it if needed
        df["_source_sheet"] = sheet

        per_sheet[sheet] = len(df)
        frames.append(df)

    if not frames:
        return pd.DataFrame(), sheets, per_sheet

    combined = pd.concat(frames, ignore_index=True, sort=False)
    return combined, sheets, per_sheet


# ---------------------------------------------------------------------------
# Type mapping
# ---------------------------------------------------------------------------
TYPE_MAP = {
    "MARCHANDISES DIVERS": [
        "COLIS","ENGIN","CABINES","CAISSES","LEGER","LOURD",
        "UTILITAIRE","PALLETES","CITERNE","UNITES","PACKAGES",
    ],
    "MINERAIS ET PRODUITS METALLURGIQUES": [
        "TUBES","CORNIERES","FIL MACHINE","ROND A BETON","BOBINES",
        "POUTRELLES","CHARPENTE METALIQUE","COUDES","ELINGUES",
        "TOLES EN PLAQUES","BILLETTES","FER PLAT","RAILS","TOLES",
        "BARRES","IRON ORE","PELLETS","HBI","CDRI","BITUME",
        "CONTENEUR","ENGINES","MDF","PLYWOOD","BOIS",
    ],
    "PRODUITS AGRICOLES / DENREES ALIMENTAIRES": [
        "BLE","HUILE DE SOJA","AVOINE","MAIS","HARICOT",
    ],
    "MINERAUX ET MATERIAUX DE CONSTRUCTION": [
        "ANHYDRITE","ARGILE","BALL CLAY","CIMENT","CLINKERS",
        "FELDSPAR","KAOLIN","MARBRE","SABLE SILICEUX","BOIS BLANC","BOIS ROUGE",
    ],
    "ENGRAIS ET PRODUITS CHIMIQUES": [
        "SULFATE","SULFATE DE SOUDE","SODIUM","PVC","PVC EDGE BANDING",
    ],
    "PRODUITS PETROLIERS": ["BITUME"],
}

_KEYWORD_TO_TYPE: dict[str, str] = {}
for _cat, _keywords in TYPE_MAP.items():
    for _kw in _keywords:
        _KEYWORD_TO_TYPE[_kw.upper()] = _cat


# ---------------------------------------------------------------------------
# Core logic helpers  (updated to accept col_map)
# ---------------------------------------------------------------------------
def gs_type_function(item: str) -> str:
    if not isinstance(item, str):
        return "MARCHANDISES DIVERS"
    item_upper = item.strip().upper()
    if item_upper in _KEYWORD_TO_TYPE:
        return _KEYWORD_TO_TYPE[item_upper]
    best_match, best_cat = "", "MARCHANDISES DIVERS"
    for kw, cat in _KEYWORD_TO_TYPE.items():
        if kw in item_upper and len(kw) > len(best_match):
            best_match, best_cat = kw, cat
    return best_cat


def gs_extract_chassis(val) -> str:
    if pd.isna(val) or not str(val).strip():
        return ""
    s = re.sub(r"\s+", "", str(val).strip().upper())
    if not (8 <= len(s) <= 17):
        return ""
    if not (any(c.isalpha() for c in s) and any(c.isdigit() for c in s)):
        return ""
    if not re.match(r"^[A-Z0-9]+$", s):
        return ""
    return s


def gs_has_valid_chassis(row) -> bool:
    """Works on rows that already have canonical column names."""
    raw = row.get("Némuro de chassis", "")
    if pd.isna(raw):
        raw = ""
    return bool(gs_extract_chassis(raw))


def gs_marchandise_d(row):
    val = row.get("Marchandise", "")
    return "" if pd.isna(val) else val


def gs_marchandise_h(row):
    val = row.get("Marchandise.1", "")
    return "" if pd.isna(val) else val


def gs_normalize_number(val, as_int=False):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    try:
        num = float(str(val).replace(",", ".").replace(" ", ""))
        return int(round(num)) if as_int else num
    except (ValueError, TypeError):
        return ""


def gs_filter_nombre_colis(df: pd.DataFrame):
    if "nombre colis" not in df.columns:
        return df.copy(), 0
    mask = df["nombre colis"].apply(
        lambda x: not pd.isna(x) and str(x).strip() != ""
    )
    filtered = df[mask].copy().reset_index(drop=True)
    return filtered, len(df) - len(filtered)


def gs_filter_valid_chassis(df: pd.DataFrame):
    if "Némuro de chassis" not in df.columns:
        return pd.DataFrame(columns=df.columns), len(df)
    mask = df.apply(gs_has_valid_chassis, axis=1)
    filtered = df[mask].copy().reset_index(drop=True)
    return filtered, len(df) - len(filtered)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
DATA_FONT    = Font(name="Calibri", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN         = Side(border_style="thin", color="B8CCE4")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT1_FILL    = PatternFill("solid", fgColor="DEEAF1")
ALT2_FILL    = PatternFill("solid", fgColor="FFFFFF")
DATE_FMT     = "DD/MM/YYYY"
NUM_FMT      = "#,##0.00"
INT_FMT      = "#,##0"


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment, cell.border = ALIGN_CENTER, BORDER


def _style_data(ws, row, ncols, alt):
    fill = ALT1_FILL if alt else ALT2_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = fill, DATA_FONT
        cell.alignment, cell.border = ALIGN_LEFT, BORDER


def _auto_width(ws, mn=10, mx=40):
    for col_cells in ws.columns:
        ml = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max(ml + 2, mn), mx)


def _freeze(ws, row=2):
    ws.freeze_panes = ws.cell(row=row, column=1)


# ---------------------------------------------------------------------------
# Sheet builders  (now receive a DataFrame with canonical column names)
# ---------------------------------------------------------------------------
def gs_build_navires(ws, df: pd.DataFrame, append: bool, cancel_flag: threading.Event):
    headers = [
        "N° ESCALE","NAVIRE","CONSIGNATAIRE","QUAI","DATE","EMBARQ/DEBARQ",
        "TYPES","PRODUITS","DETAIL","QUANTITE","TONNAGE","CLIENTS",
        "DATE FINITION OPERATION","FIN ENLEVEMENT","JOURS","ENTREPOSAGE",
        "SURFACE","BL / TRANSITAIRE","ADRESSE","N° CHASSIS",
    ]

    df_f, skipped = gs_filter_nombre_colis(df)

    if not append:
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30
        start_row = 2
    else:
        start_row = ws.max_row + 1

    chassis_list = [
        gs_extract_chassis(df_f.iloc[i].get("Némuro de chassis", "") or "")
        for i in range(len(df_f))
    ]

    for idx in range(len(df_f)):
        if cancel_flag.is_set():
            return None, None

        er  = start_row + idx
        row = df_f.iloc[idx]
        pv  = str(row.get("PRODUITS", "") or "")

        values = [
            row.get("escale_numb", ""),
            row.get("navire_name", ""),
            row.get("CONSIGNATAIRE", ""),
            "",
            row.get("DATE", ""),
            row.get("EMBARQ/DEBARQ", ""),
            gs_type_function(pv),
            pv,
            gs_marchandise_d(row),
            gs_normalize_number(row.get("nombre colis", ""), True),
            gs_normalize_number(row.get("Poids brute", "")),
            row.get("Client", ""),
            None,
            TODAY,
            f'=IFERROR(N{er}-M{er},"")',
            "TP",
            gs_normalize_number(row.get("SURF", "")),
            row.get("BL", ""),
            "",
            chassis_list[idx],
        ]

        for ci, v in enumerate(values, 1):
            ws.cell(row=er, column=ci, value=v)

        _style_data(ws, er, len(headers), alt=(idx % 2 == 0))
        ws.row_dimensions[er].height = 18
        ws.cell(er,  5).number_format = DATE_FMT
        ws.cell(er, 13).number_format = DATE_FMT
        ws.cell(er, 14).number_format = DATE_FMT
        ws.cell(er, 10).number_format = INT_FMT
        ws.cell(er, 11).number_format = NUM_FMT
        ws.cell(er, 17).number_format = NUM_FMT

    if not append:
        _auto_width(ws)
        _freeze(ws, 2)

    return len(df_f), skipped


def gs_build_chassis(ws, df: pd.DataFrame, append: bool, cancel_flag: threading.Event):
    headers = [
        "NAVIRE","N° BL","Article","Marchandise","PRODUITS",
        "nombre colis","Poids brute","Client","Marchandise",
        "Némuro de chassis","MODEL","DESIGNATION",
    ]
    hints = [
        "","","","(col D source)","(Détails PRODUITS)","","","",
        "(col H source)","(extracted)","","(col D source)",
    ]

    df_f, skipped = gs_filter_valid_chassis(df)

    if not append:
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30
        for ci, hint in enumerate(hints, 1):
            cell = ws.cell(row=2, column=ci, value=hint)
            cell.fill      = SUBHDR_FILL
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=9, italic=True)
            cell.alignment = ALIGN_CENTER
            cell.border    = BORDER
        ws.row_dimensions[2].height = 14
        start_row = 3
    else:
        start_row = ws.max_row + 1

    for idx in range(len(df_f)):
        if cancel_flag.is_set():
            return None, None

        er  = start_row + idx
        row = df_f.iloc[idx]

        values = [
            row.get("navire_name", ""),
            row.get("BL", ""),
            row.get("Article", ""),
            gs_marchandise_d(row),
            row.get("Détails PRODUITS", ""),
            gs_normalize_number(row.get("nombre colis", ""), True),
            gs_normalize_number(row.get("Poids brute", "")),
            row.get("Client", ""),
            gs_marchandise_h(row),
            row.get("Némuro de chassis", ""),
            row.get("Modèle", ""),
            gs_marchandise_d(row),
        ]

        for ci, v in enumerate(values, 1):
            ws.cell(row=er, column=ci, value=v)

        _style_data(ws, er, len(headers), alt=(idx % 2 == 0))
        ws.row_dimensions[er].height = 18
        ws.cell(er, 6).number_format = INT_FMT
        ws.cell(er, 7).number_format = NUM_FMT

    if not append:
        _auto_width(ws)
        _freeze(ws, 3)

    return len(df_f), skipped


def gs_build_rst(ws, df: pd.DataFrame, append: bool, cancel_flag: threading.Event):
    headers = [
        "N° ESCALE","NAVIRE","DATE D'ENTREE","DATE FINITION",
        "QUANTITE MANIFETE","TONNAGE MANIFETE","SURFACE MANIFETE","CLIENT",
        "Détails PRODUITS","PRODUITS","08 J","2 M 08 J","N° GROS / MRN",
        "RESTE QUANTITE","RESTE TONNAGE","RESTE SURFACE","B/L",
        "CRN / ART","DESIGNATION","LATITUDE","LONGITUDE","ETAT",
    ]

    df_f, skipped = gs_filter_nombre_colis(df)

    if not append:
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        _style_header(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30
        start_row = 2
    else:
        start_row = ws.max_row + 1

    for idx in range(len(df_f)):
        if cancel_flag.is_set():
            return None, None

        er  = start_row + idx
        row = df_f.iloc[idx]

        qte     = gs_normalize_number(row.get("nombre colis", ""), True)
        tonnage = gs_normalize_number(row.get("Poids brute", ""))
        surface = gs_normalize_number(row.get("SURF", ""))

        values = [
            row.get("escale_numb", ""),
            row.get("navire_name", ""),
            row.get("DATE D'ENTREE", ""),
            "",
            qte, tonnage, surface,
            row.get("Client", ""),
            row.get("Détails PRODUITS", ""),
            row.get("PRODUITS", ""),
            f'=IFERROR(D{er}+8,"")',
            f'=IFERROR(K{er}+8,"")',
            row.get("MRN", ""),
            qte, tonnage, surface,
            row.get("BL", ""),
            row.get("Article", ""),
            gs_marchandise_d(row),
            "", "",
            "N-OP",
        ]

        for ci, v in enumerate(values, 1):
            ws.cell(row=er, column=ci, value=v)

        _style_data(ws, er, len(headers), alt=(idx % 2 == 0))
        ws.row_dimensions[er].height = 18
        for c in [3, 4, 11, 12]:
            ws.cell(er, c).number_format = DATE_FMT
        for c in [5, 14]:
            ws.cell(er, c).number_format = INT_FMT
        for c in [6, 7, 15, 16]:
            ws.cell(er, c).number_format = NUM_FMT

    if not append:
        _auto_width(ws)
        _freeze(ws, 2)

    return len(df_f), skipped


# ---------------------------------------------------------------------------
# Main generator (runs in background thread)
# ---------------------------------------------------------------------------
def gs_run_generation(
    raw: pd.DataFrame,
    cancel_flag: threading.Event,
    result: dict,
):
    """
    `raw` already has canonical column names (produced by load_all_sheets).
    Builds output_omar.xlsx on disk and stores result dict.
    """
    try:
        if os.path.exists(OUTPUT_FILE):
            wb = load_workbook(OUTPUT_FILE)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        stats = {}

        # ── NAVIRES ────────────────────────────────────────────────────────
        sn  = "NAVIRES"
        app = sn in wb.sheetnames
        ws  = wb[sn] if app else wb.create_sheet(sn)
        if not app:
            ws.sheet_view.showGridLines = False
        kept, skipped = gs_build_navires(ws, raw, app, cancel_flag)
        if cancel_flag.is_set():
            result["cancelled"] = True
            return
        stats["NAVIRES"] = {"kept": kept, "skipped": skipped, "appended": app}

        # ── N° chassis ─────────────────────────────────────────────────────
        sn  = "N° chassis"
        app = sn in wb.sheetnames
        ws  = wb[sn] if app else wb.create_sheet(sn)
        if not app:
            ws.sheet_view.showGridLines = False
        kept, skipped = gs_build_chassis(ws, raw, app, cancel_flag)
        if cancel_flag.is_set():
            result["cancelled"] = True
            return
        stats["N° chassis"] = {"kept": kept, "skipped": skipped, "appended": app}

        # ── RST ────────────────────────────────────────────────────────────
        sn  = RST_SHEET_NAME
        app = sn in wb.sheetnames
        ws  = wb[sn] if app else wb.create_sheet(sn)
        if not app:
            ws.sheet_view.showGridLines = False
        kept, skipped = gs_build_rst(ws, raw, app, cancel_flag)
        if cancel_flag.is_set():
            result["cancelled"] = True
            return
        stats[RST_SHEET_NAME] = {"kept": kept, "skipped": skipped, "appended": app}

        wb.save(OUTPUT_FILE)
        result["stats"] = stats

    except Exception as exc:
        result["error"] = str(exc)


# ---------------------------------------------------------------------------
# Streamlit page
# ---------------------------------------------------------------------------
def page_generate_sheets():
    st.title("📋 Generate Sheets")
    st.caption(
        f"Upload a source file to generate and update **`{OUTPUT_FILE}`**."
    )

    # ── File uploader ──────────────────────────────────────────────────────
    uploaded_file = st.file_uploader(
        "📂 Upload Source Manifest File (Excel)", type=["xlsx", "xls"]
    )
    if uploaded_file is not None:
        with open(SOURCE_FILE, "wb") as f:
            f.write(uploaded_file.getbuffer())
        # Reset reverse map so it's rebuilt fresh on next resolve
        global _REVERSE_MAP
        _REVERSE_MAP = {}

    # ── Session state init ─────────────────────────────────────────────────
    for key, default in [
        ("gs_running",     False),
        ("gs_cancel_flag", None),
        ("gs_thread",      None),
        ("gs_result",      {}),
        ("gs_done",        False),
        ("gs_raw",         None),       # canonical DataFrame cached here
        ("gs_sheet_info",  {}),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── Source file presence check ─────────────────────────────────────────
    source_exists = os.path.exists(SOURCE_FILE)
    output_exists = os.path.exists(OUTPUT_FILE)

    c1, c2 = st.columns(2)
    with c1:
        if source_exists:
            st.success(f"✅ Source ready: `{SOURCE_FILE}`")
        else:
            st.error("❌ Please upload a source file above.")
    with c2:
        if output_exists:
            st.info(f"📂 Output exists: `{OUTPUT_FILE}` — will **append**")
        else:
            st.info(f"🆕 Output not found — will **create** `{OUTPUT_FILE}`")

    st.divider()

    # ── Load & resolve all sheets ──────────────────────────────────────────
    raw = None
    if source_exists:
        try:
            raw, sheet_names, per_sheet = load_all_sheets(SOURCE_FILE)
            st.session_state.gs_raw        = raw
            st.session_state.gs_sheet_info = per_sheet
        except Exception as e:
            st.error(f"❌ Cannot read `{SOURCE_FILE}`: {e}")
            return

        # ── Sheet breakdown ────────────────────────────────────────────────
        with st.expander(f"📑 Sheets found in source ({len(sheet_names)})", expanded=True):
            sheet_df = pd.DataFrame(
                [{"Sheet": s, "Rows": per_sheet.get(s, 0)} for s in sheet_names]
            )
            st.dataframe(sheet_df, use_container_width=True, hide_index=True)

        # ── Metrics ───────────────────────────────────────────────────────
        total_rows = len(raw)
        nc_filled  = (
            raw["nombre colis"]
            .apply(lambda x: not pd.isna(x) and str(x).strip() != "")
            .sum()
            if "nombre colis" in raw.columns else 0
        )
        chassis_ok = (
            raw.apply(gs_has_valid_chassis, axis=1).sum()
            if "Némuro de chassis" in raw.columns else 0
        )
        bl_count   = (
            raw["BL"].nunique() if "BL" in raw.columns else 0
        )

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Rows (all sheets)", total_rows)
        m2.metric("→ NAVIRES / RST rows",    nc_filled)
        m3.metric("→ N° chassis rows",       chassis_ok)
        m4.metric("Unique BL",               bl_count)

        # ── Column resolution report ───────────────────────────────────────
        required_canonicals = [
            "BL", "navire_name", "nombre colis",
            "Poids brute", "PRODUITS", "escale_numb",
        ]
        missing = [c for c in required_canonicals if c not in raw.columns]
        if missing:
            st.warning(
                f"⚠️ Could not resolve required columns: "
                f"`{'`, `'.join(missing)}`  — check aliases or add them to COLUMN_ALIASES."
            )

        with st.expander("🗺️ Column resolution map (canonical → actual)"):
            # Show per-sheet maps for transparency
            xl = pd.ExcelFile(SOURCE_FILE)
            for sname in xl.sheet_names:
                try:
                    df_tmp = xl.parse(sname, header=0, dtype=str)
                    df_tmp.columns = [str(c).strip() for c in df_tmp.columns]
                    cmap   = build_column_map(df_tmp.columns)
                    cmap_df = pd.DataFrame(
                        [{"Canonical": k, "Actual column": v} for k, v in cmap.items()]
                    )
                    st.markdown(f"**Sheet: {sname}**")
                    st.dataframe(cmap_df, use_container_width=True, hide_index=True)
                except Exception:
                    pass

        with st.expander("🔍 Preview combined data (first 10 rows)"):
            st.dataframe(raw.head(10), use_container_width=True)

    st.divider()

    # ── Poll running thread ────────────────────────────────────────────────
    if st.session_state.gs_running:
        thread: threading.Thread = st.session_state.gs_thread
        if thread and not thread.is_alive():
            st.session_state.gs_running = False
            st.session_state.gs_done    = True

    # ── Control buttons ────────────────────────────────────────────────────
    st.subheader("⚙️ Controls")
    btn_col1, btn_col2 = st.columns([1, 1])

    with btn_col1:
        start_disabled = (
            not source_exists
            or raw is None
            or st.session_state.gs_running
        )
        if st.button(
            "▶ Start Processing",
            type="primary",
            disabled=start_disabled,
            use_container_width=True,
        ):
            st.session_state.gs_done   = False
            st.session_state.gs_result = {}
            cancel_flag                = threading.Event()
            result_dict: dict          = {}
            st.session_state.gs_cancel_flag = cancel_flag
            st.session_state.gs_result      = result_dict

            t = threading.Thread(
                target=gs_run_generation,
                args=(st.session_state.gs_raw, cancel_flag, result_dict),
                daemon=True,
            )
            st.session_state.gs_thread  = t
            st.session_state.gs_running = True
            t.start()
            st.rerun()

    with btn_col2:
        stop_disabled = not st.session_state.gs_running
        if st.button(
            "⏹ Stop / Cancel",
            type="secondary",
            disabled=stop_disabled,
            use_container_width=True,
        ):
            if st.session_state.gs_cancel_flag:
                st.session_state.gs_cancel_flag.set()
            st.warning("⚠️ Cancellation requested — stopping after current row …")

    st.divider()

    # ── Live status while running ──────────────────────────────────────────
    if st.session_state.gs_running:
        with st.spinner("⏳ Processing … click **Stop / Cancel** to abort."):
            import time
            time.sleep(1)
            st.rerun()

    # ── Result display ─────────────────────────────────────────────────────
    if st.session_state.gs_done:
        result = st.session_state.gs_result

        if result.get("cancelled"):
            st.warning("🚫 Processing was cancelled. Output file was **not** saved.")

        elif result.get("error"):
            st.error(f"❌ Error during processing:\n\n`{result['error']}`")

        elif result.get("stats"):
            st.success("✅ Processing complete!")

            st.subheader("📊 Sheet Summary")
            rows_data = []
            for sheet, s in result["stats"].items():
                rows_data.append({
                    "Sheet":        sheet,
                    "Mode":         "Appended ➕" if s["appended"] else "Created 🆕",
                    "Rows Written": s["kept"],
                    "Rows Skipped": s["skipped"],
                })
            st.dataframe(
                pd.DataFrame(rows_data),
                use_container_width=True,
                hide_index=True,
            )

            st.divider()
            if os.path.exists(OUTPUT_FILE):
                with open(OUTPUT_FILE, "rb") as f:
                    file_bytes = f.read()
                st.download_button(
                    label=f"⬇️ Download  {OUTPUT_FILE}",
                    data=file_bytes,
                    file_name=OUTPUT_FILE,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            else:
                st.error("Output file not found on disk after processing.")

    # ── How-it-works info ─────────────────────────────────────────────────
    if not st.session_state.gs_running and not st.session_state.gs_done:
        with st.expander("ℹ️ How it works"):
            st.markdown(f"""
            **Multi-sheet loading**: Every sheet in `{SOURCE_FILE}` is read and stacked.

            **Fuzzy header resolution** (4-pass strategy per sheet):
            | Pass | Method |
            |------|--------|
            | 1 | Exact match |
            | 2 | Case-insensitive + normalised punctuation |
            | 3 | Token-subset match (word order irrelevant) |
            | 4 | SequenceMatcher fuzzy ratio ≥ 0.82 |

            | Output Sheet | Filter Rule |
            |-------------|-------------|
            | **NAVIRES** | `nombre colis` is filled |
            | **N° chassis** | `Némuro de chassis` is valid (8-17 alphanumeric) |
            | **{RST_SHEET_NAME}** | `nombre colis` is filled |

            - Output: **`{OUTPUT_FILE}`** (appended if it already exists)
            """)
