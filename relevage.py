import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os


def find_rst_sheet(workbook):
    """
    Find the sheet matching pattern 'RST DD-MM-YY' or similar date formats
    using regex. Returns the sheet name or None.
    """
    # Pattern: RST followed by a date-like string
    # Matches: RST 23-03-26 M, RST 24-01-15, RST 23-12-01 X, etc.
    pattern = re.compile(r'^RST\s+\d{2}-\d{2}-\d{2}', re.IGNORECASE)

    matched_sheets = []
    for sheet_name in workbook.sheetnames:
        if pattern.search(sheet_name.strip()):
            matched_sheets.append(sheet_name)

    if len(matched_sheets) == 0:
        return None
    elif len(matched_sheets) == 1:
        return matched_sheets[0]
    else:
        # Multiple matches → return the last one (most recent)
        return matched_sheets[-1]


def update_suivi(data_suivi_path, update_suivi_path, output_path, progress_callback=None):
    """
    Update the data_suivi Excel file based on removal records in update_suivi.
    """

    def log(msg):
        print(msg)
        if progress_callback:
            progress_callback(msg)

    # ============================================================
    # STEP 1: Clean update_suivi - Remove duplicate N°Enlevement
    # ============================================================
    log("Step 1: Cleaning update_suivi (removing duplicate N°Enlevement)...")

    df_update = pd.read_excel(update_suivi_path)
    log(f"  Original update rows: {len(df_update)}")

    df_update = df_update.drop_duplicates(subset=['N°Enlevement'], keep='first')
    log(f"  After removing duplicates: {len(df_update)}")

    # ============================================================
    # STEP 2: Load data_suivi workbook & find RST sheet via regex
    # ============================================================
    log("Step 2: Loading data_suivi workbook...")

    wb = openpyxl.load_workbook(data_suivi_path)

    log(f"  Available sheets: {wb.sheetnames}")

    rst_sheet_name = find_rst_sheet(wb)

    if rst_sheet_name is None:
        raise ValueError(
            f"No sheet matching 'RST DD-MM-DD ...' found.\n"
            f"Available sheets: {wb.sheetnames}"
        )

    log(f"  Found RST sheet: '{rst_sheet_name}'")
    ws = wb[rst_sheet_name]

    # ============================================================
    # STEP 3: Identify column indices using regex on headers
    # ============================================================
    log("Step 3: Identifying column headers with regex...")

    headers = {}
    header_row = 1

    for row in range(1, 10):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and re.search(r'ESCALE', str(cell_value), re.IGNORECASE):
            header_row = row
            break

    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if cell_value:
            headers[str(cell_value).strip()] = col

    log(f"  Header row: {header_row}")

    # Regex patterns for each column
    col_escale = None
    col_quantite_manifeste = None
    col_reste_quantite = None
    col_reste_surface = None
    col_bl = None
    col_etat = None

    for h, c in headers.items():
        if re.match(r'^N[°˚º]?\s*ESCALE$', h, re.IGNORECASE):
            col_escale = c
        elif re.match(r'^QUANTIT[EÉ]\s*MANIFES?T[EÉ]E?$', h, re.IGNORECASE):
            col_quantite_manifeste = c
        elif re.match(r'^RESTE\s*QUANTIT[EÉ]$', h, re.IGNORECASE):
            col_reste_quantite = c
        elif re.match(r'^RESTE\s*SURFACE$', h, re.IGNORECASE):
            col_reste_surface = c
        elif re.match(r'^B\s*/?\s*L$', h, re.IGNORECASE):
            col_bl = c
        elif re.match(r'^[EÉ]TAT$', h, re.IGNORECASE):
            col_etat = c

    log(f"  N° ESCALE:         col {col_escale} ({get_column_letter(col_escale) if col_escale else '?'})")
    log(f"  QUANTITE MANIFETE: col {col_quantite_manifeste} ({get_column_letter(col_quantite_manifeste) if col_quantite_manifeste else '?'})")
    log(f"  RESTE QUANTITE:    col {col_reste_quantite} ({get_column_letter(col_reste_quantite) if col_reste_quantite else '?'})")
    log(f"  RESTE SURFACE:     col {col_reste_surface} ({get_column_letter(col_reste_surface) if col_reste_surface else '?'})")
    log(f"  B/L:               col {col_bl} ({get_column_letter(col_bl) if col_bl else '?'})")
    log(f"  ETAT:              col {col_etat} ({get_column_letter(col_etat) if col_etat else '?'})")

    if None in [col_escale, col_quantite_manifeste, col_reste_quantite, col_reste_surface, col_bl, col_etat]:
        missing = []
        if col_escale is None: missing.append("N° ESCALE")
        if col_quantite_manifeste is None: missing.append("QUANTITE MANIFETE")
        if col_reste_quantite is None: missing.append("RESTE QUANTITE")
        if col_reste_surface is None: missing.append("RESTE SURFACE")
        if col_bl is None: missing.append("B/L")
        if col_etat is None: missing.append("ETAT")
        raise ValueError(f"Could not find columns: {missing}")

    # ============================================================
    # STEP 4: Group updates by (Escale, BL)
    # ============================================================
    log("Step 4: Grouping updates by (Escale, BL)...")

    updates_grouped = {}
    for _, row in df_update.iterrows():
        escale = str(row['Escale']).strip()
        bl = str(row['N° BL']).strip()
        colis_enlev = row['Colis Enlev']

        try:
            colis_enlev = float(colis_enlev)
        except (ValueError, TypeError):
            log(f"  WARNING: Skipping non-numeric Colis Enlev: {colis_enlev}")
            continue

        key = (escale, bl)
        if key not in updates_grouped:
            updates_grouped[key] = []
        updates_grouped[key].append(colis_enlev)

    log(f"  Total unique (Escale, BL) groups: {len(updates_grouped)}")

    # ============================================================
    # STEP 5: Define fill styles
    # ============================================================
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    turquoise_fill = PatternFill(start_color="40E0D0", end_color="40E0D0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ============================================================
    # STEP 6: Process each row
    # ============================================================
    log("\nStep 6: Processing rows...")

    col_letter_qte = get_column_letter(col_quantite_manifeste)
    rows_updated = 0
    rows_soldee = 0
    rows_oper = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        escale_cell = ws.cell(row=row_idx, column=col_escale).value
        bl_cell = ws.cell(row=row_idx, column=col_bl).value

        if escale_cell is None and bl_cell is None:
            continue

        escale_val = str(escale_cell).strip() if escale_cell else ""
        bl_val = str(bl_cell).strip() if bl_cell else ""

        key = (escale_val, bl_val)

        if key in updates_grouped:
            removals = updates_grouped[key]

            # Get existing content
            existing_value = ws.cell(row=row_idx, column=col_reste_quantite).value
            existing_str = str(existing_value).strip() if existing_value else ""

            log(f"\n  Row {row_idx}: Escale={escale_val}, BL={bl_val}")
            log(f"    Existing RESTE QUANTITE: {existing_str}")
            log(f"    New removals: {removals}")

            # Build formula — preserve existing if starts with "="
            if re.match(r'^=', existing_str):
                formula = existing_str
            else:
                formula = f"={col_letter_qte}{row_idx}"

            for removal_val in removals:
                if removal_val == int(removal_val):
                    formula += f"-{int(removal_val)}"
                else:
                    formula += f"-{removal_val}"

            log(f"    Final formula: {formula}")

            ws.cell(row=row_idx, column=col_reste_quantite).value = formula

            # Calculate remaining quantity
            quantite_manifeste = ws.cell(row=row_idx, column=col_quantite_manifeste).value
            try:
                qte_manifeste_num = float(quantite_manifeste) if quantite_manifeste else 0
            except (ValueError, TypeError):
                qte_manifeste_num = 0

            all_subtractions = re.findall(r'-\s*([\d.]+)', formula)
            total_subtracted = sum(float(val) for val in all_subtractions)
            reste_quantite_calc = qte_manifeste_num - total_subtracted

            log(f"    QTE Manifeste: {qte_manifeste_num}, Subtracted: {total_subtracted}, RESTE: {reste_quantite_calc}")

            # ============================================================
            # FORMATTING
            # ============================================================

            if reste_quantite_calc <= 0:
                # SOLDEE: Col A → Col ETAT (X) → YELLOW + RED text
                log(f"    → SOLDEE")

                for col in range(col_escale, col_etat + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = yellow_fill
                    cell.font = Font(
                        color="FF0000",
                        bold=cell.font.bold if cell.font else False,
                        size=cell.font.size if cell.font else None,
                        name=cell.font.name if cell.font else None
                    )

                ws.cell(row=row_idx, column=col_etat).value = "SOLDEE"
                ws.cell(row=row_idx, column=col_etat).fill = yellow_fill
                ws.cell(row=row_idx, column=col_etat).font = Font(color="FF0000", bold=True)
                rows_soldee += 1

            elif reste_quantite_calc > 0 and reste_quantite_calc != qte_manifeste_num:
                # OPER: Col A → Col P → LIGHT BLUE 80%, ETAT → TURQUOISE
                log(f"    → OPER")

                for col in range(col_escale, col_reste_surface + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = light_blue_fill

                ws.cell(row=row_idx, column=col_etat).value = "OPER"
                ws.cell(row=row_idx, column=col_etat).fill = turquoise_fill
                rows_oper += 1

            else:
                log(f"    → No change (RESTE = QTE MANIFETE)")

            rows_updated += 1

    # ============================================================
    # STEP 7: Save
    # ============================================================
    log(f"\nStep 7: Saving to: {output_path}")
    wb.save(output_path)

    summary = (
        f"\n{'=' * 50}\n"
        f"SUMMARY\n"
        f"{'=' * 50}\n"
        f"  RST Sheet used:      {rst_sheet_name}\n"
        f"  Total rows updated:  {rows_updated}\n"
        f"  Rows marked OPER:    {rows_oper}\n"
        f"  Rows marked SOLDEE:  {rows_soldee}\n"
        f"  Output: {output_path}\n"
        f"{'=' * 50}"
    )
    log(summary)

    return rows_updated, rows_oper, rows_soldee, rst_sheet_name


# ============================================================
# SIMPLE UI
# ============================================================

class SuiviUpdaterApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Suivi Updater")

        # 1. Define window dimensions
        window_width = 650
        window_height = 400

        # 2. Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # 3. Calculate center coordinates
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height - window_height/1.5)

        # 4. Set geometry: "width x height + x_offset + y_offset"
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

        self.root.resizable(True, True)

        # Variables
        self.data_suivi_path = tk.StringVar()
        self.update_suivi_path = tk.StringVar()
        self.build_ui()

    def build_ui(self):
        # ---- Title ----
        title_frame = tk.Frame(self.root, bg="#2C3E50", height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(
            title_frame,
            text="📦 Suivi Updater — Mise à jour des restes",
            font=("Segoe UI", 14, "bold"),
            bg="#2C3E50",
            fg="white"
        ).pack(pady=15)

        # ---- File Selection Frame ----
        file_frame = tk.LabelFrame(
            self.root,
            text="Sélection des fichiers",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=15
        )
        file_frame.pack(fill=tk.X, padx=15, pady=(15, 5))

        # Data Suivi
        tk.Label(
            file_frame,
            text="📁 Fichier Data Suivi (workbook with RST sheet):",
            font=("Segoe UI", 9),
            anchor="w"
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 2))

        data_entry = tk.Entry(
            file_frame,
            textvariable=self.data_suivi_path,
            width=60,
            font=("Segoe UI", 9)
        )
        data_entry.grid(row=1, column=0, sticky="ew", padx=(0, 5))

        tk.Button(
            file_frame,
            text="Parcourir...",
            command=self.browse_data_suivi,
            font=("Segoe UI", 9),
            bg="#3498DB",
            fg="white",
            relief="flat",
            padx=10
        ).grid(row=1, column=1, sticky="e")

        # Update Suivi
        tk.Label(
            file_frame,
            text="📁 Fichier Update Suivi (enlèvements):",
            font=("Segoe UI", 9),
            anchor="w"
        ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(15, 2))

        update_entry = tk.Entry(
            file_frame,
            textvariable=self.update_suivi_path,
            width=60,
            font=("Segoe UI", 9)
        )
        update_entry.grid(row=3, column=0, sticky="ew", padx=(0, 5))

        tk.Button(
            file_frame,
            text="Parcourir...",
            command=self.browse_update_suivi,
            font=("Segoe UI", 9),
            bg="#3498DB",
            fg="white",
            relief="flat",
            padx=10
        ).grid(row=3, column=1, sticky="e")

        file_frame.columnconfigure(0, weight=1)

        # ---- Run Button ----
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(fill=tk.X, padx=15, pady=10)

        self.run_button = tk.Button(
            btn_frame,
            text="▶  Lancer la mise à jour",
            command=self.run_update,
            font=("Segoe UI", 11, "bold"),
            bg="#27AE60",
            fg="white",
            relief="flat",
            padx=30,
            pady=8,
            cursor="hand2"
        )
        self.run_button.pack(fill=tk.X)

        # ---- Log Frame ----
        log_frame = tk.LabelFrame(
            self.root,
            text="Journal d'exécution",
            font=("Segoe UI", 10, "bold"),
            padx=10,
            pady=10
        )
        log_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 15))

        self.log_text = tk.Text(
            log_frame,
            height=15,
            font=("Consolas", 9),
            bg="#1E1E1E",
            fg="#00FF00",
            insertbackground="white",
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        scrollbar = tk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(fill=tk.Y, side=tk.RIGHT)
        self.log_text.config(yscrollcommand=scrollbar.set)

    def browse_data_suivi(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier Data Suivi",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*")
            ]
        )
        if path:
            self.data_suivi_path.set(path)

    def browse_update_suivi(self):
        path = filedialog.askopenfilename(
            title="Sélectionner le fichier Update Suivi",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*")
            ]
        )
        if path:
            self.update_suivi_path.set(path)

    def log_message(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def run_update(self):
        data_path = self.data_suivi_path.get().strip()
        update_path = self.update_suivi_path.get().strip()

        # Validate inputs
        if not data_path:
            messagebox.showerror("Erreur", "Veuillez sélectionner le fichier Data Suivi.")
            return
        if not update_path:
            messagebox.showerror("Erreur", "Veuillez sélectionner le fichier Update Suivi.")
            return
        if not os.path.exists(data_path):
            messagebox.showerror("Erreur", f"Fichier introuvable:\n{data_path}")
            return
        if not os.path.exists(update_path):
            messagebox.showerror("Erreur", f"Fichier introuvable:\n{update_path}")
            return

        # Build output path
        base, ext = os.path.splitext(data_path)
        output_path = f"{base}_updated{ext}"

        # Clear log
        self.log_text.delete("1.0", tk.END)

        # Disable button during processing
        self.run_button.config(state=tk.DISABLED, text="⏳ Traitement en cours...", bg="#95A5A6")
        self.root.update_idletasks()

        try:
            rows_updated, rows_oper, rows_soldee, rst_sheet = update_suivi(
                data_path,
                update_path,
                output_path,
                progress_callback=self.log_message
            )

            messagebox.showinfo(
                "Terminé ✅",
                f"Mise à jour terminée avec succès!\n\n"
                f"Feuille RST: {rst_sheet}\n"
                f"Lignes mises à jour: {rows_updated}\n"
                f"OPER: {rows_oper}\n"
                f"SOLDEE: {rows_soldee}\n\n"
                f"Fichier sauvegardé:\n{output_path}"
            )

        except Exception as e:
            self.log_message(f"\n❌ ERREUR: {str(e)}")
            messagebox.showerror("Erreur", f"Une erreur est survenue:\n\n{str(e)}")

        finally:
            self.run_button.config(state=tk.NORMAL, text="▶  Lancer la mise à jour", bg="#27AE60")

    def run(self):
        self.root.mainloop()


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    app = SuiviUpdaterApp()
    app.run()